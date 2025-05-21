from django.shortcuts import render
from django.http import FileResponse, HttpResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os, ast
from django.conf import settings
from io import BytesIO
from rest_framework.decorators import api_view
from rest_framework.response import Response
from . serializers import MachineSrlzr
from rest_framework import status
from .models import MachineData

"""
AWS ROOT PASSWORD: {PC Password} + 01
"""

def returnHomePage(requests):
    return HttpResponse("<div><h1>WELCOME TO THE HOMEPAGE FOR SAM MACHINE OUTPUT MONITORING</h1></div>")

def handleExcel(data):
    try:
        excel_path = os.path.join(settings.BASE_DIR, "staticfiles", "TEMPLATE.xlsx")
        wb = load_workbook(excel_path)
        ws =  wb.active

        ws["G1"].value = data["name"]
        ws["M1"].value = data["date"]
        #b5 b6 b8 b9
        ws["B5"].value = data["ds_ok"]
        ws["B6"].value = data["ds_ng"]
        ws["B8"].value = data["ns_ok"]
        ws["B9"].value = data["ns_ng"]

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        print("Success")

        return output
    except:
        print("An error occured")


# def generate_excel(request):
#     data = request.GET.get('value')

#     res = handleExcel(data)

#     return FileResponse(res, as_attachment=True, filename="TestFile.xlsx")

@api_view(["POST"])
def setData(request):
    try:
        res = request.data
        date = request.data.get("date")
        shift = request.data.get("shift")
        machine = request.data.get("machine")
        obj, create = MachineData.objects.update_or_create(
            date=date,
            shift=shift,
            machine=machine,
            defaults=res
        )
        srlzr = MachineSrlzr(obj)
        
        return Response(srlzr.data, status=status.HTTP_200_OK)
    except:
        return Response({"Message": "Failed"}, status=status.HTTP_400_BAD_REQUEST)

@api_view(["GET"])
def getData(request, machine):
    data = MachineData.objects.get(machine=machine)

    srlzr = MachineSrlzr(data)
    
    try:
        sam_dict = {
            "name": f"SAM {srlzr.data['machine']}",
            "date": srlzr.data["date"],
            "ds_ok": srlzr.data["ds_ok_count"],
            "ds_ng": srlzr.data["ds_ng_count"],
            "ns_ok": srlzr.data["ns_ok_count"],
            "ns_ng": srlzr.data["ns_ng_count"],
            "ds_ok_hr": srlzr.data["ds_ok_perhr"],
            "ds_ng_hr": srlzr.data["ds_ng_perhr"],
            "ns_ok_hr": srlzr.data["ns_ok_perhr"],
            "ns_ng_hr": srlzr.data["ns_ng_perhr"]
        }

        res = handleExcel(sam_dict)
        
        return FileResponse(res, as_attachment=True, filename=f"{sam_dict['name']}.xlsx")
    
    except:
        return Response({"Message": "OUTPUT FAILED, Try again later"}, status=status.HTTP_404_NOT_FOUND)

@api_view(["DELETE"])
def deleteData(request, machine):
    machine = MachineData.objects.filter(machine=machine)
    try:
        machine.delete()

        return Response({"Message": "Machine succesfully deleted"}, status=status.HTTP_200_OK)
    except:
        return Response({"Message": "Failed to delete machine"})

def handleExcelForAll(data):
    try:
        path = os.path.join(settings.BASE_DIR, "staticfiles", "TEMPLATE.xlsx")
        wb = load_workbook(path)
        ws = wb["OUTPUT"]

        ws.sheet_state = "veryHidden"
        for item in data:
            ws2 = wb.copy_worksheet(ws)
            ws2["G1"].value = item["name"]
            ws2["M1"].value = item["date"]
            #b5 b6 b8 b9
            ws2["B5"].value = item["ds_ok"]
            ws2["B6"].value = item["ds_ng"]
            ws2["B8"].value = item["ns_ok"]
            ws2["B9"].value = item["ns_ng"]

            for ndx, obj in enumerate(item['ds_ok_hr']):
                col = get_column_letter(ndx + 4)
                ws2[f"{col}5"].value = obj
            for ndx, obj in enumerate(item['ds_ng_hr']):
                col = get_column_letter(ndx + 4)
                ws2[f"{col}6"].value = obj
            for ndx, obj in enumerate(item['ns_ok_hr']):
                col = get_column_letter(ndx + 4)
                ws2[f"{col}8"].value = obj
            for ndx, obj in enumerate(item['ns_ng_hr']):
                col = get_column_letter(ndx + 4)
                ws2[f"{col}9"].value = obj



            ws2.title = item["name"]

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
    except: pass
 

@api_view(["GET"])
def getAllData(request, date):
    try:
        data = MachineData.objects.filter(date=date)

        srlzr = MachineSrlzr(data, many=True)

        data_list = []

        data_arr = srlzr.data
        for item in data_arr:
            data_to_send = {
                "name": f"SAM {item['machine']}",
                "date": item["date"],
                "shift": item["shift"],
                "ds_ok": item["ds_ok_count"],
                "ds_ng": item["ds_ng_count"],
                "ns_ok": item["ns_ok_count"],
                "ns_ng": item["ns_ng_count"],
                "ds_ok_hr": ast.literal_eval(item["ds_ok_perhr"]),
                "ds_ng_hr": ast.literal_eval(item["ds_ng_perhr"]),
                "ns_ok_hr": ast.literal_eval(item["ns_ok_perhr"]),
                "ns_ng_hr": ast.literal_eval(item["ns_ng_perhr"])
            }

            data_list.append(data_to_send)

        res = handleExcelForAll(data_list)
        
        return FileResponse(res, as_attachment=True, filename=f"SAM OUTPUT {date}.xlsx")
    except Exception as e:
        return HttpResponse("""
                            <div>
                                <h1>No data found, please try other dates</h1>
                            </div>
                            """)

@api_view(["GET"])
def getJSON(requests):
    data = MachineData.objects.all()

    data_srlzr = MachineSrlzr(data, many=True)

    try:
        return Response(data_srlzr.data, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"Message": "It failed for some reason"})
